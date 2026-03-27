"""Rebuild Excel files with 7-commodity, 27-district structure."""
import io, sys, warnings
warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = r'c:/Users/Administrator/Documents/BI/sarima and arima'
np.random.seed(42)

# ─────────────────────────────────────────────────────────────────────────────
# MASTER DATA
# ─────────────────────────────────────────────────────────────────────────────
DISTRICT_MASTER = [
    ('Kampala','Central',1.13),('Natete','Central',1.08),
    ('Luwero','Central',0.96),('Kayunga','Central',0.94),('Gomba','Central',0.90),
    ('Mubende','Western',0.96),('Hoima','Western',0.99),('Masindi','Western',0.97),
    ('Kibaale','Western',0.91),('Kyegegwa','Western',0.92),
    ('Kasese','Western',0.95),('Mutukula','Western',1.04),
    ('Gulu','Northern',0.92),('Lira','Northern',0.90),('Kiryadongo','Northern',0.87),
    ('Bweyale','Northern',0.85),('Nwoya','Northern',0.87),('Alebtong','Northern',0.84),
    ('Pader','Northern',0.83),('Kitgum','Northern',0.82),
    ('Mbale','Eastern',0.98),('Jinja','Eastern',1.04),
    ('Busia','Eastern',0.96),('Soroti','Eastern',0.93),
    ('Kigezi','Western',1.05),('Kapchorwa','Eastern',1.02),('Kabale','Western',1.00),
]
ALL_DISTRICTS    = [d[0] for d in DISTRICT_MASTER]
DIST_REGION      = {d[0]: d[1] for d in DISTRICT_MASTER}
BARLEY_DISTRICTS = ['Kigezi','Kapchorwa','Kabale']
MAIN_DISTRICTS   = [d for d in ALL_DISTRICTS if d not in BARLEY_DISTRICTS]
MONTHS           = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

COMMODITIES = ['Maize','Sorghum White','Sorghum Red',
               'Beans Yellow','Beans Nambaale','Beans Wairimu','Barley','Soya']
COMMODITY_DISTRICTS = {c: (BARLEY_DISTRICTS if c=='Barley' else MAIN_DISTRICTS)
                        for c in COMMODITIES}

# Colour scheme
DARK='0A1520'; ALT_D='0D1F2D'; ALT_M='122030'; LITE='C8D8E8'; GOLD='FFD700'
RCOL  = {'Central':'E91E8C','Western':'9C27B0','Northern':'FF5722','Eastern':'00BCD4'}
CCOL  = {'Maize':'1E88E5','Sorghum White':'66BB6A','Sorghum Red':'EF5350',
          'Beans Yellow':'FFA726','Beans Nambaale':'AB47BC',
          'Beans Wairimu':'26C6DA','Barley':'8D6E63','Soya':'8BC34A'}

def hf(h): return PatternFill('solid', fgColor=h)
def mf(bold=False,color='C8D8E8',size=10):
    return Font(bold=bold,color=color,size=size,name='Calibri')
def ctr(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def lft(): return Alignment(horizontal='left',vertical='center')

# ─────────────────────────────────────────────────────────────────────────────
# LOAD CSVs
# ─────────────────────────────────────────────────────────────────────────────
df_long      = pd.read_csv(f'{BASE}/PBI_Uganda_Grains_Daily.csv')
df_long['Date'] = pd.to_datetime(df_long['Date'], dayfirst=True)
df_monthly   = pd.read_csv(f'{BASE}/PBI_Uganda_Grains_Monthly.csv')
df_monthly['Date'] = pd.to_datetime(df_monthly['Date'], dayfirst=False)
df_forecasts = pd.read_csv(f'{BASE}/PBI_Uganda_Grains_Forecasts.csv')
df_forecasts['Date'] = pd.to_datetime(df_forecasts['Date'], dayfirst=False)
df_stats     = pd.read_csv(f'{BASE}/PBI_Model_Statistics.csv')
actuals      = df_long[df_long['Data_Type']=='Actual'].copy()
print(f"Loaded: {len(df_long):,} rows, commodities: {actuals['Commodity'].unique().tolist()}")

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WIDE FRAMES (for daily sheets)
# ─────────────────────────────────────────────────────────────────────────────
print("Building wide frames...")
df_wide = {}
for year in [2024,2025]:
    sub_yr = actuals[actuals['Date'].dt.year==year]
    dates  = sorted(sub_yr['Date'].unique())
    rows   = []
    for d in dates:
        sub_d = sub_yr[sub_yr['Date']==d]
        row   = {'Date':pd.Timestamp(d),'Day':pd.Timestamp(d).strftime('%a')}
        for commodity in COMMODITIES:
            for district in COMMODITY_DISTRICTS[commodity]:
                mask = (sub_d['Commodity']==commodity)&(sub_d['District']==district)
                row[f'{commodity}_{district}'] = (
                    int(sub_d.loc[mask,'Price_UGX'].values[0]) if mask.any() else '')
        rows.append(row)
    df_wide[year] = pd.DataFrame(rows)
    print(f"  {year}: {len(rows)} rows")

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
print("Building workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── README ────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('README')
readme_lines = [
    ('Uganda Grain Commodity Price Intelligence', True, GOLD, 14),
    ('District-Level Daily Prices 2024-2025  |  UGX/kg', False, LITE, 11),
    ('', False, LITE, 10),
    ('COMMODITIES (8)', True, GOLD, 11),
    ('Maize              -- All 24 general districts', False, LITE, 10),
    ('Sorghum White      -- All 24 districts (Northern/Eastern harvest cycle)', False, LITE, 10),
    ('Sorghum Red        -- All 24 districts (Eastern/Central harvest cycle)', False, LITE, 10),
    ('Beans Yellow       -- All 24 districts (two harvest seasons: Jul-Aug + Dec-Jan)', False, LITE, 10),
    ('Beans Nambaale     -- All 24 districts (specialty variety, higher price)', False, LITE, 10),
    ('Beans Wairimu      -- All 24 districts (mixed/other varieties)', False, LITE, 10),
    ('Barley             -- ONLY Kigezi, Kapchorwa, Kabale (SW and Mt Elgon highlands)', False, LITE, 10),
    ('Soya               -- All 24 districts (Northern/Eastern production; lean Jul-Sep)', False, LITE, 10),
    ('', False, LITE, 10),
    ('DISTRICT MARKETS (27 total: 24 general + 3 barley-only highland)', True, GOLD, 11),
] + [
    (f'{d[0]} ({d[1]}) -- {"BARLEY ONLY" if d[0] in BARLEY_DISTRICTS else "all crops"}', False, LITE, 10)
    for d in DISTRICT_MASTER
] + [
    ('', False, LITE, 10),
    ('ADDITIONAL SOURCE ENTRIES', True, GOLD, 11),
    ('Grain Pulse -- Kampala broker index (see notes)', False, LITE, 10),
    ('Agroways -- trader benchmark', False, LITE, 10),
    ('Nairobi (KSh) -- cross-border ref at ~28 UGX/KSh', False, LITE, 10),
    ('Natete -- merged into Kampala district', False, LITE, 10),
    ('Bweyale -- merged into Kiryadongo district', False, LITE, 10),
    ('Mutukula -- included as separate Western district', False, LITE, 10),
    ('', False, LITE, 10),
    ('Sources: FEWS NET, WFP VAM, Agahikaine Grains Ltd, FAO GIEWS', False, '8A9BAB', 9),
]
for ri, (txt,bold,color,size) in enumerate(readme_lines,start=1):
    c = ws.cell(row=ri,column=2,value=txt)
    c.font = Font(bold=bold,color=color,size=size,name='Calibri')
    c.fill = hf(DARK)
ws.column_dimensions['B'].width = 72
ws.sheet_view.showGridLines = False
print("  README done")

# ── Daily 2024 / Daily 2025 ───────────────────────────────────────────────────
for year in [2024,2025]:
    ws = wb.create_sheet(f'Daily {year}')
    # Each commodity has its own group of columns
    # Layout: Date | Day | [Maize x 24] | [Sorg White x 24] | ... | [Barley x 3]
    col_groups = []
    for commodity in COMMODITIES:
        districts = COMMODITY_DISTRICTS[commodity]
        col_groups.append((commodity, districts))

    total_cols = 2 + sum(len(d) for _, d in col_groups)

    # Row 1: title
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws['A1']
    c.value = (f'Uganda Grain Prices -- Daily District Markets {year}  |  UGX/kg  '
               f'|  8 Commodities  |  27 Markets (24 general + 3 barley-only)')
    c.font = mf(True,GOLD,12); c.fill = hf(DARK); c.alignment = ctr()
    ws.row_dimensions[1].height = 24

    # Row 2: sources
    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    c = ws['A2']
    c.value = 'Sources: FEWS NET, WFP VAM, Agahikaine Grains Ltd, FAO GIEWS  |  CONFIDENTIAL'
    c.font = mf(False,'8A9BAB',9); c.fill = hf(DARK); c.alignment = lft()

    # Row 3: commodity group headers
    ws.cell(row=3,column=1).fill = hf(DARK)
    ws.cell(row=3,column=2).fill = hf(DARK)
    col_cursor = 3
    for commodity, districts in col_groups:
        n = len(districts)
        col_e = col_cursor + n - 1
        ws.merge_cells(f'{get_column_letter(col_cursor)}3:{get_column_letter(col_e)}3')
        c = ws.cell(row=3,column=col_cursor,value=f'{commodity.upper()}  (UGX/kg)')
        c.font = mf(True,CCOL.get(commodity,'FFFFFF'),10)
        c.fill = hf(DARK); c.alignment = ctr()
        col_cursor += n
    ws.row_dimensions[3].height = 18

    # Row 4: district name headers
    for col,val in [(1,'Date'),(2,'Day')]:
        c = ws.cell(row=4,column=col,value=val)
        c.font = mf(True,GOLD); c.fill = hf(DARK); c.alignment = ctr()
    col_cursor = 3
    for commodity, districts in col_groups:
        for district in districts:
            region = DIST_REGION[district]
            c = ws.cell(row=4,column=col_cursor,value=district)
            c.font = mf(True,RCOL.get(region,'FFFFFF'),8)
            c.fill = hf(DARK); c.alignment = ctr()
            col_cursor += 1
    ws.row_dimensions[4].height = 28

    # Row 5: region sub-labels
    ws.cell(row=5,column=1,value='Region:').font = mf(False,'8A9BAB',8)
    ws.cell(row=5,column=1).fill = hf(ALT_D)
    ws.cell(row=5,column=2).fill = hf(ALT_D)
    col_cursor = 3
    for commodity, districts in col_groups:
        for district in districts:
            region = DIST_REGION[district]
            c = ws.cell(row=5,column=col_cursor,value=region)
            c.font = mf(False,RCOL.get(region,'AAAAAA'),7)
            c.fill = hf(ALT_D); c.alignment = ctr()
            col_cursor += 1
    ws.row_dimensions[5].height = 12

    # Column widths
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 5
    for col in range(3, total_cols+1):
        ws.column_dimensions[get_column_letter(col)].width = 8

    # Data rows
    for ri,(_, row) in enumerate(df_wide[year].iterrows()):
        rn = ri+6
        fc = ALT_D if ri%2==0 else ALT_M
        dt = row['Date']
        c = ws.cell(row=rn,column=1,value=dt.strftime('%d-%b-%Y'))
        c.font = mf(False,LITE,9); c.fill = hf(fc); c.alignment = ctr()
        c = ws.cell(row=rn,column=2,value=dt.strftime('%a'))
        c.font = mf(False,'8A9BAB',9); c.fill = hf(fc); c.alignment = ctr()
        col_cursor = 3
        for commodity, districts in col_groups:
            for district in districts:
                price = row.get(f'{commodity}_{district}','')
                c = ws.cell(row=rn,column=col_cursor,value=price)
                c.font = mf(False,LITE,9); c.fill = hf(fc); c.alignment = ctr()
                col_cursor += 1
    ws.freeze_panes = 'A6'
    print(f"  Daily {year}: {len(df_wide[year])} rows, {total_cols} cols")

# ── Monthly Summary ───────────────────────────────────────────────────────────
ws = wb.create_sheet('Monthly Summary')
hdr = ['Commodity','District','Region','Year']+MONTHS+['Annual Avg']
nc  = len(hdr)
ws.merge_cells(f'A1:{get_column_letter(nc)}1')
c = ws['A1']; c.value = 'Uganda Grain Prices -- Monthly District Averages (UGX/kg)'
c.font = mf(True,GOLD,12); c.fill = hf(DARK); c.alignment = lft()
ws.merge_cells(f'A2:{get_column_letter(nc)}2')
c = ws['A2']; c.value = ('Barley: Kigezi, Kapchorwa, Kabale only  |  '
                          'Beans: two-season calendar (Jul-Aug + Dec-Jan harvest)  |  '
                          'Sorghum White/Red shown separately  |  '
                          'Soya: all 24 districts, harvest Oct-Dec')
c.font = mf(False,'8A9BAB',9); c.fill = hf(DARK)
for ci,h in enumerate(hdr,start=1):
    c = ws.cell(row=3,column=ci,value=h)
    c.font = mf(True,GOLD); c.fill = hf(DARK); c.alignment = ctr()
dr = 4
for commodity in COMMODITIES:
    for district in COMMODITY_DISTRICTS[commodity]:
        region = DIST_REGION[district]
        for year in [2024,2025]:
            fc = ALT_D if dr%2==0 else ALT_M
            sub = actuals[(actuals['Commodity']==commodity) &
                          (actuals['District']==district) &
                          (actuals['Date'].dt.year==year)]
            ma = sub.groupby(sub['Date'].dt.month)['Price_UGX'].mean()
            rv = [commodity,district,region,year]
            rv += [int(ma.get(m,np.nan)) if m in ma.index and not np.isnan(ma.get(m,np.nan)) else '' for m in range(1,13)]
            rv.append(int(sub['Price_UGX'].mean()) if len(sub)>0 else '')
            for ci,val in enumerate(rv,start=1):
                c = ws.cell(row=dr,column=ci,value=val)
                c.font = mf(False,LITE,9); c.fill = hf(fc); c.alignment = ctr()
            dr += 1
ws.freeze_panes = 'A4'
for ci,w in enumerate([14,12,10,6]+[8]*13,start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
print(f"  Monthly Summary: {dr-4} rows")

# ── Forecasts 2026 ────────────────────────────────────────────────────────────
ws = wb.create_sheet('Forecasts 2026')
hdr_f = ['Commodity','District','Region','Model']+MONTHS+['Annual Avg']
nf    = len(hdr_f)
ws.merge_cells(f'A1:{get_column_letter(nf)}1')
c = ws['A1']; c.value = 'Uganda Grain Price Forecasts 2026  |  District-Level  |  UGX/kg'
c.font = mf(True,GOLD,12); c.fill = hf(DARK); c.alignment = lft()
ws.merge_cells(f'A2:{get_column_letter(nf)}2')
c = ws['A2']; c.value = ('SARIMA(1,1,1)(1,1,1)12 + SARIMAX regional demand  |  '
                           'Blended = 40% SARIMA + 60% SARIMAX')
c.font = mf(False,'8A9BAB',9); c.fill = hf(DARK)
for ci,h in enumerate(hdr_f,start=1):
    c = ws.cell(row=3,column=ci,value=h)
    c.font = mf(True,GOLD); c.fill = hf(DARK); c.alignment = ctr()
dr = 4
for commodity in COMMODITIES:
    for district in COMMODITY_DISTRICTS[commodity]:
        region = DIST_REGION[district]
        sf = df_forecasts[(df_forecasts['Commodity']==commodity) &
                          (df_forecasts['District']==district)]
        for model,col_name in [('SARIMA','SARIMA_Forecast'),
                                ('SARIMAX','SARIMAX_Forecast'),
                                ('Blended','Blended_Forecast'),
                                ('Lower 95% CI','Lower_95CI'),
                                ('Upper 95% CI','Upper_95CI')]:
            fc = ALT_D if dr%2==0 else ALT_M
            rv = [commodity,district,region,model]
            for m in range(1,13):
                r = sf[sf['Month']==m]
                v = r[col_name].values[0] if len(r)>0 else np.nan
                rv.append(int(v) if not pd.isna(v) else '')
            ann = sf[col_name].mean()
            rv.append(int(ann) if not pd.isna(ann) else '')
            for ci,val in enumerate(rv,start=1):
                c = ws.cell(row=dr,column=ci,value=val)
                c.font = mf(False,LITE,9); c.fill = hf(fc); c.alignment = ctr()
            dr += 1
ws.freeze_panes = 'A4'
for ci,w in enumerate([14,12,10,16]+[8]*13,start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
print(f"  Forecasts 2026: {dr-4} rows")

# ── Model Statistics ──────────────────────────────────────────────────────────
ws = wb.create_sheet('Model Statistics')
hdr_s = ['Commodity','District','Region','N Obs','Mean 2024','Mean 2025',
         'YoY Delta','Trend/mth','Seas Range %','MAE','MAPE %']
ws.merge_cells(f'A1:{get_column_letter(len(hdr_s))}1')
c = ws['A1']; c.value = 'SARIMA/SARIMAX Model Performance -- District-Level Summary'
c.font = mf(True,GOLD,12); c.fill = hf(DARK)
for ci,h in enumerate(hdr_s,start=1):
    c = ws.cell(row=2,column=ci,value=h)
    c.font = mf(True,GOLD); c.fill = hf(DARK); c.alignment = ctr()
for ri,(_, row) in enumerate(df_stats.iterrows(),start=3):
    fc = ALT_D if ri%2==0 else ALT_M
    yoy = round(float(row['Mean_Price_2025'])-float(row['Mean_Price_2024']),0)
    rv  = [row['Commodity'],row['District'],row['Region'],row['N_obs'],
           int(row['Mean_Price_2024']),int(row['Mean_Price_2025']),int(yoy),
           row['Trend_UGX_per_month'],row['Seasonality_Range_Pct'],
           int(row['MAE']),row['MAPE_Pct']]
    for ci,val in enumerate(rv,start=1):
        c = ws.cell(row=ri,column=ci,value=val)
        c.font = mf(False,LITE,9); c.fill = hf(fc); c.alignment = ctr()
ws.freeze_panes = 'A3'
for ci,w in enumerate([14,12,10,7,10,10,10,9,11,8,8],start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
print(f"  Model Statistics: {len(df_stats)} rows")

# ── Seasonal Analysis ─────────────────────────────────────────────────────────
ws = wb.create_sheet('Seasonal Analysis')
nc_sa = 16
ws.merge_cells(f'A1:{get_column_letter(nc_sa)}1')
c = ws['A1']; c.value = 'Seasonal Price Index -- District Markets (Base = Annual Average = 100)'
c.font = mf(True,GOLD,12); c.fill = hf(DARK)
ws.merge_cells(f'A2:{get_column_letter(nc_sa)}2')
c = ws['A2']; c.value = ('Index >110 = Lean Season premium (red)  |  '
                          'Index <90 = Post-Harvest discount (green)  |  '
                          'Beans: dual harvest seasons (Jul-Aug + Dec-Jan)')
c.font = mf(False,'8A9BAB',9); c.fill = hf(DARK)
hdr_sa = ['Commodity','District','Region','Year']+MONTHS
for ci,h in enumerate(hdr_sa,start=1):
    c = ws.cell(row=3,column=ci,value=h)
    c.font = mf(True,GOLD); c.fill = hf(DARK); c.alignment = ctr()

def sifill(val):
    if pd.isna(val) or val=='': return hf(ALT_D)
    if val>=118: return hf('B71C1C')
    if val>=110: return hf('E53935')
    if val>=105: return hf('EF6C00')
    if val<=80:  return hf('1B5E20')
    if val<=90:  return hf('2E7D32')
    if val<=95:  return hf('388E3C')
    return hf(ALT_D)

dr = 4
for commodity in COMMODITIES:
    for district in COMMODITY_DISTRICTS[commodity]:
        region = DIST_REGION[district]
        for year in [2024,2025]:
            sub = actuals[(actuals['Commodity']==commodity) &
                          (actuals['District']==district) &
                          (actuals['Date'].dt.year==year)]
            ma  = sub.groupby(sub['Date'].dt.month)['Price_UGX'].mean()
            ann = ma.mean()
            rv  = [commodity,district,region,year]
            for m in range(1,13):
                v  = ma.get(m,np.nan)
                si = round(v/ann*100,1) if not pd.isna(v) and ann>0 else ''
                rv.append(si)
            for ci,val in enumerate(rv,start=1):
                c = ws.cell(row=dr,column=ci,value=val)
                if ci>4 and val!='':
                    c.fill = sifill(val); c.font = mf(False,'FFFFFF',9)
                else:
                    c.fill = hf(ALT_D if dr%2==0 else ALT_M)
                    c.font = mf(False,LITE,9)
                c.alignment = ctr()
            dr += 1
ws.freeze_panes = 'A4'
for ci,w in enumerate([14,12,10,6]+[8]*12,start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
print(f"  Seasonal Analysis: {dr-4} rows")

# ── API Sources Guide (preserve) ─────────────────────────────────────────────
ws_api = wb.create_sheet('API Sources Guide')
try:
    wb_src = openpyxl.load_workbook(f'{BASE}/Uganda_Grain_Prices_Daily_2024_2025.xlsx', data_only=True)
    if 'API Sources Guide' in wb_src.sheetnames:
        for row in wb_src['API Sources Guide'].iter_rows():
            for cell in row:
                ws_api.cell(row=cell.row,column=cell.column,value=cell.value)
        print("  API Sources Guide: copied")
    else:
        ws_api.cell(row=1,column=2,value='API Sources Guide (see separate documentation)')
except Exception as e:
    print(f"  API Sources Guide: skipped ({e})")

wb.save(f'{BASE}/Uganda_Grain_Prices_Daily_2024_2025.xlsx')
print(f"Saved Uganda_Grain_Prices_Daily_2024_2025.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# UPDATE PowerBI_Dashboard_Guide.xlsx
# ─────────────────────────────────────────────────────────────────────────────
print("Updating PowerBI_Dashboard_Guide.xlsx ...")
wb_pbi = openpyxl.load_workbook(f'{BASE}/PowerBI_Dashboard_Guide.xlsx')
ws_dm  = wb_pbi['Data Model']

max_r = ws_dm.max_row + 3
c = ws_dm.cell(row=max_r,column=2,value='COMMODITY & DISTRICT STRUCTURE (UPDATED)')
c.font = Font(bold=True,size=12,color='1565C0',name='Calibri')

# Commodity summary
for ci,h in enumerate(['Commodity','Districts Covered','Base Price 2024','Notes'],start=2):
    ws_dm.cell(row=max_r+1,column=ci,value=h).font = Font(bold=True,size=10,name='Calibri')

commodity_notes = {
    'Maize':          (len(MAIN_DISTRICTS), '~870 UGX/kg', 'Two harvests Aug-Sep + Jan-Feb'),
    'Sorghum White':  (len(MAIN_DISTRICTS), '~1,050 UGX/kg','Northern harvest Oct-Dec'),
    'Sorghum Red':    (len(MAIN_DISTRICTS), '~980 UGX/kg', 'Eastern harvest Aug-Sep'),
    'Beans Yellow':   (len(MAIN_DISTRICTS), '~2,400 UGX/kg','Dual season Jul-Aug + Dec-Jan'),
    'Beans Nambaale': (len(MAIN_DISTRICTS), '~2,650 UGX/kg','Specialty; Central/Eastern focus'),
    'Beans Wairimu':  (len(MAIN_DISTRICTS), '~2,200 UGX/kg','Mixed varieties; widespread'),
    'Barley':         (3,                  '~1,550 UGX/kg','ONLY Kigezi, Kapchorwa, Kabale'),
    'Soya':           (len(MAIN_DISTRICTS), '~1,950 UGX/kg','All 24 districts; harvest Oct-Dec'),
}
for i,(commodity,(n_d,price,note)) in enumerate(commodity_notes.items(),start=max_r+2):
    for ci,val in enumerate([commodity,f'{n_d} districts',price,note],start=2):
        c = ws_dm.cell(row=i,column=ci,value=val)
        c.font = Font(size=9,name='Calibri',
                      color='BF360C' if commodity=='Barley' else '1A237E')

wb_pbi.save(f'{BASE}/PowerBI_Dashboard_Guide.xlsx')
print("Saved PowerBI_Dashboard_Guide.xlsx")
print()
print("ALL DONE")
